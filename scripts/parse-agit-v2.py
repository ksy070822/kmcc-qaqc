#!/usr/bin/env python3
"""
아지트 품질관리 데이터 통합 파싱 v2

데이터 소스: 아지트글취합/ (광주 + 용산, 25.7~26.2)
  - 품질 아지트.xlsx  (용산/광주 시트 → agit_posts, qc_weekly_actions, education_records)
  - 광주/용산 폴더 내 부진상담사 엑셀 → underperforming_weekly

출력: /tmp/agit_bq/*.ndjson → bq load --replace
"""

import os
import re
import json
import hashlib
import unicodedata
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

import zipfile
import tempfile
import shutil

import openpyxl

# ============================================================
# 설정
# ============================================================
PROJECT_ID = "csopp-25f2"
DATASET_ID = "KMCC_QC"
BASE_DIR = Path("/Users/may.08/Desktop/AI 자동화 구축 모음/kmcc-qc-dashbord")
AGIT_DIR = BASE_DIR / "아지트글취합 "   # 끝에 공백 있음
XLSX_FILE = AGIT_DIR / "품질 아지트.xlsx"
OUT_DIR = Path("/tmp/agit_bq")
OUT_DIR.mkdir(exist_ok=True)

# QC 16개 항목 매핑
QC_ITEM_MAP = {
    "첫인사": ("greeting_error", "상담태도"),
    "끝인사": ("greeting_error", "상담태도"),
    "첫인사/끝인사": ("greeting_error", "상담태도"),
    "첫인사끝인사": ("greeting_error", "상담태도"),
    "공감표현": ("empathy_error", "상담태도"),
    "사과표현": ("apology_error", "상담태도"),
    "추가문의": ("additional_inquiry_error", "상담태도"),
    "불친절": ("unkind_error", "상담태도"),
    "상담유형": ("consult_type_error", "오상담/오처리"),
    "상담유형 오설정": ("consult_type_error", "오상담/오처리"),
    "가이드": ("guide_error", "오상담/오처리"),
    "가이드 미준수": ("guide_error", "오상담/오처리"),
    "본인확인": ("identity_check_error", "오상담/오처리"),
    "필수탐색": ("required_search_error", "오상담/오처리"),
    "오안내": ("wrong_guide_error", "오상담/오처리"),
    "전산처리누락": ("process_missing_error", "오상담/오처리"),
    "전산 처리 누락": ("process_missing_error", "오상담/오처리"),
    "전산처리미흡": ("process_incomplete_error", "오상담/오처리"),
    "전산 처리 미흡": ("process_incomplete_error", "오상담/오처리"),
    "전산조작": ("system_error", "오상담/오처리"),
    "전산 조작": ("system_error", "오상담/오처리"),
    "콜픽트립": ("id_mapping_error", "오상담/오처리"),
    "ID매핑": ("id_mapping_error", "오상담/오처리"),
    "플래그": ("flag_keyword_error", "오상담/오처리"),
    "키워드": ("flag_keyword_error", "오상담/오처리"),
    "플래그/키워드": ("flag_keyword_error", "오상담/오처리"),
    "상담이력": ("history_error", "오상담/오처리"),
    "상담이력 기재": ("history_error", "오상담/오처리"),
}

# ============================================================
# 유틸리티
# ============================================================
def make_id(*parts):
    raw = "_".join(str(p) for p in parts)
    return hashlib.md5(raw.encode()).hexdigest()[:16]

def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def safe_float(val):
    if val is None:
        return None
    try:
        s = str(val).replace("%", "").strip()
        if s in ("", "-", "N/A", "n/a"):
            return None
        return float(s)
    except:
        return None

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.strptime(s, fmt).date()
        except:
            continue
    return None

def classify_post(title):
    if not title:
        return "other"
    t = title.lower()
    if "qc" in t and ("리포트" in t or "리포팅" in t or "report" in t):
        return "qc_weekly"
    if "보수교육" in t or "보수 교육" in t:
        return "refresher"
    if "신입" in t and "교육" in t:
        return "new_hire"
    if "직무" in t and ("테스트" in t or "평가" in t):
        return "job_test"
    if "qa" in t and ("평가" in t or "대상" in t):
        return "qa_eval"
    if "심화" in t and "교육" in t:
        return "advanced_training"
    return "other"

def map_qc_item(item_text):
    if not item_text:
        return None, "미분류"
    clean = item_text.strip()
    for key, (col, cat) in QC_ITEM_MAP.items():
        if key in clean:
            return col, cat
    return None, "미분류"

def parse_vertical(vertical):
    v = vertical.lower().replace(" ", "")
    service = "기타"
    channel = "기타"
    if "주차" in v or "카오너" in v:
        service = "주차"
    elif "바이크" in v:
        service = "바이크"
    elif "maas" in v or "택시" in v or "심야" in v:
        service = "택시"
    elif "퀵" in v:
        service = "퀵"
    elif "대리" in v:
        service = "대리"
    elif "화물" in v:
        service = "화물"
    if "유선" in v:
        channel = "유선"
    elif "채팅" in v:
        channel = "채팅"
    return service, channel

def write_ndjson(filepath, records):
    with open(filepath, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False, default=str) + "\n")
    print(f"  -> {filepath}: {len(records)}건")

def clean_xlsx_autofilter(filepath):
    """openpyxl autoFilter XML 오류 우회"""
    tmpdir = tempfile.mkdtemp()
    clean_path = os.path.join(tmpdir, 'clean.xlsx')
    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(clean_path, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                    content = data.decode('utf-8')
                    content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                    content = re.sub(r'<autoFilter[^/]*/>', '', content)
                    data = content.encode('utf-8')
                zout.writestr(item, data)
    return clean_path, tmpdir


# ============================================================
# Phase 1: agit_posts
# ============================================================
def parse_agit_posts():
    print("\n" + "=" * 60)
    print("Phase 1: agit_posts")
    print("=" * 60)

    wb = openpyxl.load_workbook(str(XLSX_FILE), read_only=True, data_only=True)
    rows = []
    now = datetime.utcnow().isoformat()

    for sheet_name in ["용산", "광주"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        center = sheet_name
        count = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            vals = list(row) + [None] * 14
            a, b, c, d, e, f, g, h, i, j = vals[:10]

            post_type = safe_str(c)
            if post_type not in ("원글", "댓글"):
                continue

            post_id = safe_str(b) or safe_str(a)
            parent_id = safe_str(a)
            if not post_id:
                continue

            created_at = None
            if d:
                try:
                    created_at = d.isoformat() if isinstance(d, datetime) else str(d)
                except:
                    pass

            content = safe_str(f)
            content_clean = re.sub(r'\{IMAGE:\d+\}', '', content).strip()

            category = classify_post(content_clean) if post_type == "원글" else None
            attachment_url = safe_str(j) if j else None
            agit_url = safe_str(i) if i else None

            rows.append({
                "post_id": post_id,
                "parent_id": parent_id,
                "post_type": post_type,
                "center": center,
                "created_at": created_at,
                "author_id": safe_str(e),
                "content": content_clean[:60000],
                "has_attachment": bool(attachment_url),
                "attachment_url": attachment_url,
                "agit_url": agit_url,
                "category": category,
                "parsed_at": now,
            })
            count += 1

        print(f"  [{center}] {count}건")

    wb.close()
    write_ndjson(OUT_DIR / "agit_posts.ndjson", rows)
    return rows


# ============================================================
# Phase 2: QC 주간리포트 → qc_weekly_actions
# ============================================================
def split_service_sections(text):
    pattern = r'\*?\s*-\s*(택시|퀵|대리|화물|바이크|주차)\s*\*?'
    matches = list(re.finditer(pattern, text))
    if not matches:
        return [("통합", text)]
    sections = []
    for i, m in enumerate(matches):
        service = m.group(1).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        sections.append((service, text[start:end].strip()))
    return sections

def parse_items_from_section(text):
    items = []
    circled = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯"

    parts = re.split(f'([{circled}])', text)
    current_items = []
    i = 0
    while i < len(parts):
        p = parts[i].strip()
        if len(p) == 1 and p in circled:
            if i + 1 < len(parts):
                current_items.append(parts[i + 1])
                i += 2
            else:
                i += 1
        else:
            if current_items:
                current_items[-1] += " " + p
            i += 1

    if not current_items:
        cm = re.search(r'-\s*원인\s*[:：]\s*(.+?)(?=-\s*방안|$)', text, re.DOTALL)
        pm = re.search(r'-\s*방안\s*[:：]\s*(.+?)$', text, re.DOTALL)
        if cm or pm:
            items.append(("기타", cm.group(1).strip() if cm else "", pm.group(1).strip() if pm else ""))
        return items

    for item_text in current_items:
        cs = re.search(r'-\s*원인', item_text)
        item_name = item_text[:cs.start()].strip() if cs else item_text.split("\n")[0].strip()

        cm = re.search(r'-\s*원인\s*[:：]\s*(.+?)(?=-\s*방안|$)', item_text, re.DOTALL)
        cause = re.sub(r'\s+', ' ', cm.group(1).strip()) if cm else ""

        pm = re.search(r'-\s*방안\s*[:：]\s*(.+?)$', item_text, re.DOTALL)
        plan_t = re.sub(r'\s+', ' ', pm.group(1).strip()) if pm else ""

        item_name = re.sub(r'\s+', ' ', item_name).strip()
        if item_name or cause or plan_t:
            items.append((item_name, cause, plan_t))

    return items

def parse_qc_weekly_actions(agit_posts):
    print("\n" + "=" * 60)
    print("Phase 2: QC 주간리포트 → qc_weekly_actions")
    print("=" * 60)

    qc_parents = {}
    for p in agit_posts:
        if p["post_type"] == "원글" and p["category"] == "qc_weekly":
            qc_parents[p["post_id"]] = p

    print(f"  QC 원글: {len(qc_parents)}건")

    qc_comments = []
    for p in agit_posts:
        if p["post_type"] == "댓글" and p["parent_id"] in qc_parents:
            content = p["content"]
            if content and ("원인" in content or "방안" in content):
                qc_comments.append(p)

    print(f"  리포팅 댓글: {len(qc_comments)}건")

    actions = []
    for comment in qc_comments:
        parent = qc_parents[comment["parent_id"]]
        parent_content = parent.get("content", "")

        wm = re.search(r'(\d+)월\s*(\d+)주차', parent_content or "")
        if not wm:
            wm = re.search(r'(\d+)월\s*(\d+)주차', comment["content"] or "")
        week_label = f"{wm.group(1)}월 {wm.group(2)}주차" if wm else ""

        report_date = None
        if comment["created_at"]:
            try:
                dt = datetime.fromisoformat(str(comment["created_at"]).replace("Z", ""))
                report_date = dt.strftime("%Y-%m-%d")
            except:
                pass

        report_week = ""
        if report_date:
            try:
                d = datetime.strptime(report_date, "%Y-%m-%d")
                report_week = f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
            except:
                pass

        content = comment["content"]
        sections = split_service_sections(content)

        for service, section_text in sections:
            items = parse_items_from_section(section_text)
            for seq, (item_name, cause, plan_text) in enumerate(items, 1):
                _, item_cat = map_qc_item(item_name)
                actions.append({
                    "action_id": make_id(comment["post_id"], service, seq),
                    "post_id": comment["post_id"],
                    "parent_id": comment["parent_id"],
                    "center": comment["center"],
                    "report_week": report_week,
                    "report_week_label": week_label,
                    "report_date": report_date,
                    "service": service,
                    "channel": None,
                    "item_seq": seq,
                    "item_name": item_name,
                    "item_category": item_cat or "미분류",
                    "cause": cause,
                    "plan": plan_text,
                    "author_id": comment["author_id"],
                    "created_at": comment["created_at"],
                })

    print(f"  파싱된 액션: {len(actions)}건")
    write_ndjson(OUT_DIR / "qc_weekly_actions.ndjson", actions)
    return actions


# ============================================================
# Phase 3: 부진상담사 (용산 + 광주)
# ============================================================

def find_underperforming_excels(center):
    """센터별 부진상담사 엑셀 파일 수집"""
    center_dir = AGIT_DIR / center
    excel_files = []

    for root, dirs, files in os.walk(str(center_dir)):
        for f in files:
            fn = unicodedata.normalize('NFC', f)
            if not fn.endswith(".xlsx"):
                continue

            if center == "용산":
                if "미흡" not in fn and "부진" not in fn:
                    continue
            elif center == "광주":
                if "집중관리" not in fn:
                    continue

            # 파일명에서 월/주차 추출
            m = re.search(r'(\d{1,2})월\s*(\d)주차', fn)
            if m:
                month, week = int(m.group(1)), int(m.group(2))
            else:
                # 파일명에 월만 있는 경우 (광주: "집중관리 대상 현황_10월.xlsx")
                mm = re.search(r'(\d{1,2})월', fn)
                if mm:
                    month = int(mm.group(1))
                    week = 0  # 주차 없으면 0
                else:
                    continue

            excel_files.append((month, week, fn, os.path.join(root, f)))

    return excel_files


def select_month_end_files(excel_files, center):
    """각 월의 마지막 주차 파일만 선택"""
    by_month = defaultdict(list)
    for month, week, fn, fp in excel_files:
        by_month[month].append((week, fn, fp))

    selected = []
    for month in sorted(by_month):
        items = sorted(by_month[month], key=lambda x: x[0])
        last_week, last_fn, last_fp = items[-1]
        selected.append((month, last_week, last_fn, last_fp))
        print(f"    {center} {month}월 → {last_fn}")

    return selected


# --- 용산 파서 ---
def find_yongsan_week_columns(all_rows, sheet_name):
    """용산: row 2에서 주차 라벨 매칭 → QA(+0), 상담태도(+1), 오상담(+2), 상담평가주(+3), 상담평가월(+4) 컬럼
    반환: (qa_col, att_col, ops_col, eval_week_col, eval_month_col) - 0-indexed
    """
    sm = re.search(r'(\d{1,2})월\s*(\d)주차', sheet_name)
    if not sm:
        return None, None, None, None, None

    target_label = f"{sm.group(1)}월 {sm.group(2)}주차"

    if len(all_rows) < 3:
        return None, None, None, None, None

    row2 = list(all_rows[2]) if all_rows[2] else []
    base_col = None
    for ci, val in enumerate(row2):
        if val is not None and target_label in str(val):
            base_col = ci
            break

    if base_col is None:
        # fallback: 마지막 주차 라벨
        for ci, val in enumerate(row2):
            if val is not None and "주차" in str(val):
                base_col = ci

    if base_col is None:
        return None, None, None, None, None

    # 주차 그룹: QA(+0), QC태도(+1), QC오상담(+2), 상담평가주(+3), 상담평가월(+4)
    return base_col, base_col + 1, base_col + 2, base_col + 3, base_col + 4


def parse_yongsan_xlsx(filepath, filename, target_month=None):
    """용산 부진상담사 엑셀: Y/N 오류, col 22 저품질 플래그
    target_month: 해당 월 시트만 파싱 (중복 방지)
    calamine 엔진 사용 (openpyxl XML 파싱 오류 회피)
    """
    import pandas as pd

    # calamine으로 모든 시트 로드
    try:
        all_sheets = pd.read_excel(filepath, sheet_name=None, header=None, engine='calamine')
    except Exception as e:
        print(f"      [WARN] calamine 실패, openpyxl 시도: {e}")
        try:
            all_sheets = pd.read_excel(filepath, sheet_name=None, header=None, engine='openpyxl')
        except Exception as e2:
            print(f"      [ERROR] 엑셀 로드 실패: {e2}")
            return []

    records = []

    for sheet_name in all_sheets:
        if "개요" in sheet_name or "요약" in sheet_name:
            continue

        sm = re.search(r'(?:\(?\s*(\d{2})년\s*\)?\s*)?(\d{1,2})월\s*(\d)주차', sheet_name)
        if not sm:
            continue

        year_2d = int(sm.group(1)) if sm.group(1) else None
        month = int(sm.group(2))
        week = int(sm.group(3))

        # target_month 필터: 반드시 월이 일치해야 파싱 (연도 무관하게 월로 필터)
        if target_month is not None and month != target_month:
            continue

        if year_2d is not None:
            year = 2000 + year_2d
        else:
            # 연도 없는 시트 → month 기반 추정 (7~12 → 2025, 1~6 → 2026)
            year = 2025 if month >= 7 else 2026

        try:
            approx_date = datetime(year, month, min(week * 7, 28))
            report_date = approx_date.strftime("%Y-%m-%d")
            report_week = f"{year}-W{approx_date.isocalendar()[1]:02d}"
        except:
            report_date = None
            report_week = ""

        week_label = f"{month}월 {week}주차"

        # DataFrame → row 튜플 리스트 변환 (openpyxl iter_rows와 동일 구조)
        df = all_sheets[sheet_name]
        all_rows = [tuple(None if pd.isna(v) else v for v in row) for row in df.values]

        qa_col, att_col, ops_col, eval_w_col, eval_m_col = find_yongsan_week_columns(all_rows, sheet_name)

        data_start = None
        for idx, row in enumerate(all_rows):
            if row and len(row) > 1 and row[1] is not None:
                try:
                    if int(row[1]) == 1:
                        data_start = idx
                        break
                except:
                    continue

        if data_start is None:
            continue

        for row in all_rows[data_start:]:
            if not row or len(row) < 7:
                continue
            try:
                int(row[1])
            except:
                continue

            vertical = safe_str(row[2])
            name = safe_str(row[3])
            eng_id = safe_str(row[4])
            if not name or not eng_id:
                continue

            # W열(22) 저품질 상담사 여부 - 0-indexed
            is_under = False
            if len(row) > 22 and row[22] is not None:
                is_under = safe_str(row[22]).upper() == "Y"

            # ★ 저품질=Y인 상담사만 추출
            if not is_under:
                continue

            service, channel = parse_vertical(vertical)
            hire_date = parse_date(row[5])
            tenure = None
            if row[6] is not None:
                try:
                    tenure = int(float(str(row[6]).replace("개월", "")))
                except:
                    pass

            # 해당 주차 QC 오류
            att_error = None
            ops_error = None
            if att_col is not None and len(row) > att_col:
                v = safe_str(row[att_col]).upper()
                if v in ("Y", "N"):
                    att_error = v == "Y"
            if ops_col is not None and len(row) > ops_col:
                v = safe_str(row[ops_col]).upper()
                if v in ("Y", "N"):
                    ops_error = v == "Y"

            # 해당 주차 QA 적발 여부
            qa_flagged = False
            if qa_col is not None and len(row) > qa_col:
                v = safe_str(row[qa_col]).upper()
                if v == "Y":
                    qa_flagged = True

            # 해당 주차 상담평가 적발 여부 (주단위 값이 숫자>0)
            eval_flagged = False
            if eval_w_col is not None and len(row) > eval_w_col:
                v = row[eval_w_col]
                if v is not None:
                    try:
                        if float(v) > 0:
                            eval_flagged = True
                    except:
                        if safe_str(v).upper() == "Y":
                            eval_flagged = True

            # 선정 사유 도출
            reasons = []
            if qa_flagged:
                reasons.append("QA미달")
            if att_error:
                reasons.append("QC미달_태도")
            if ops_error:
                reasons.append("QC미달_상담")
            if eval_flagged:
                reasons.append("상담평점")
            underperforming_reason = ",".join(reasons) if reasons else "기타"

            # 비고
            note = None
            if len(row) > 23 and row[23] is not None:
                note = safe_str(row[23])

            records.append({
                "record_id": make_id("용산", eng_id, report_week, sheet_name),
                "center": "용산",
                "report_week": report_week,
                "report_week_label": week_label,
                "report_date": report_date,
                "agent_name": name,
                "agent_english_id": eng_id,
                "vertical": vertical,
                "service": service,
                "channel": channel,
                "hire_date": hire_date.isoformat() if hire_date else None,
                "tenure_months": tenure,
                "qc_attitude_error": att_error,
                "qc_ops_error": ops_error,
                "is_underperforming": True,
                "underperforming_reason": underperforming_reason,
                "note": note,
                "source_file": filename,
                "source_post_id": None,
            })

    return records


# --- 광주 파서 ---
def parse_gwangju_xlsx(filepath, filename, target_month=None):
    """
    광주 부진상담사 엑셀:
    - 다중 서비스 섹션 (1. 주차/바이크/택시/심야/퀵, 2. 화물, 3. 대리)
    - 숫자형 오류값 (검수건, 태도오류, 상담오류, 오류율)
    - row 6: 주차 라벨 (col 8, 15, 22, ... 7칸 간격)
    - row 7: 컬럼 헤더 (검수건, 태도오류, 상담오류, 태도오류율, 상담오류율, 저점건수, 월누적건)
    - col 97: 저품질 상담사 (Y)
    - col 98: 저품질 유형 (태도, 상담, 태도&상담)
    """
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        if "개요" in sheet_name or "요약" in sheet_name or "기준" in sheet_name:
            continue

        sm = re.search(r'(\d{1,2})월\s*(\d)주차', sheet_name)
        if not sm:
            continue

        month = int(sm.group(1))
        week = int(sm.group(2))

        # target_month 필터
        if target_month is not None and month != target_month:
            continue

        # 연도 추정
        if '2026' in filename or '26년' in filename:
            year = 2026 if month <= 6 else 2025
        else:
            year = 2025 if month >= 7 else 2026

        try:
            approx_date = datetime(year, month, min(week * 7, 28))
            report_date = approx_date.strftime("%Y-%m-%d")
            report_week = f"{year}-W{approx_date.isocalendar()[1]:02d}"
        except:
            report_date = None
            report_week = ""

        week_label = f"{month}월 {week}주차"

        ws = wb[sheet_name]

        # row 6에서 주차 라벨 위치
        target_week_label = f"{month}월 {week}주"
        target_week_col = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=6, column=c).value
            if v and target_week_label in str(v):
                target_week_col = c
                break

        # 동적으로 저품질 열 찾기 (row 6에서 "저품질 상담사" 텍스트)
        under_col = None
        note_col = None
        for c in range(1, ws.max_column + 1):
            v6 = ws.cell(row=6, column=c).value
            if v6 and "저품질 상담사" in str(v6):
                under_col = c
            if v6 and "저품질 유형" in str(v6):
                note_col = c
        # 섹션마다 NO행에도 저품질 열이 있을 수 있음
        if under_col is None:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=7, column=c).value
                if v and "저품질 상담사" in str(v):
                    under_col = c
                if v and "저품질 유형" in str(v):
                    note_col = c

        # 서비스 섹션 파싱
        section_starts = []
        for r in range(1, ws.max_row + 1):
            v3 = ws.cell(row=r, column=3).value
            if v3 and re.match(r'\d+\.', str(v3).strip()):
                section_starts.append(r)

        for si, section_row in enumerate(section_starts):
            # 각 섹션에서도 저품질 열 위치 확인 (섹션마다 다를 수 있음)
            sec_under_col = under_col
            sec_note_col = note_col
            # 섹션 NO행 (section_row + 1)에서 검색
            no_row = section_row + 1
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=no_row, column=c).value
                if v and "저품질 상담사" in str(v):
                    sec_under_col = c
                if v and "저품질 유형" in str(v):
                    sec_note_col = c

            data_start = section_row + 3
            section_end = section_starts[si + 1] if si + 1 < len(section_starts) else ws.max_row + 1

            for r in range(data_start, section_end):
                no_val = ws.cell(row=r, column=2).value
                if no_val is None:
                    continue
                try:
                    int(no_val)
                except:
                    continue

                vertical = safe_str(ws.cell(row=r, column=3).value)
                name = safe_str(ws.cell(row=r, column=4).value)
                eng_id = safe_str(ws.cell(row=r, column=5).value)
                if not name or not eng_id:
                    continue

                service, channel = parse_vertical(vertical)
                hire_date = parse_date(ws.cell(row=r, column=6).value)
                tenure = None
                t_val = ws.cell(row=r, column=7).value
                if t_val is not None:
                    try:
                        tenure = int(float(str(t_val).replace("개월", "")))
                    except:
                        pass

                # 해당 주차 QC 데이터
                att_error = None
                ops_error = None
                if target_week_col:
                    att_val = ws.cell(row=r, column=target_week_col + 1).value
                    ops_val = ws.cell(row=r, column=target_week_col + 2).value
                    if att_val is not None:
                        try:
                            att_error = int(float(str(att_val))) > 0
                        except:
                            pass
                    if ops_val is not None:
                        try:
                            ops_error = int(float(str(ops_val))) > 0
                        except:
                            pass

                # 저품질 상담사 (동적 열)
                is_under = False
                if sec_under_col:
                    v_under = ws.cell(row=r, column=sec_under_col).value
                    if v_under is not None:
                        is_under = safe_str(v_under).upper() == "Y"

                # ★ 저품질=Y인 상담사만 추출 (전체 로스터 제외)
                if not is_under:
                    continue

                # 저품질 유형 (태도, 상담, 태도&상담)
                under_type = None
                if sec_note_col:
                    v_note = ws.cell(row=r, column=sec_note_col).value
                    if v_note:
                        under_type = safe_str(v_note)

                # 선정 사유 도출: 저품질 유형 기반
                reasons = []
                if under_type:
                    if "태도" in under_type:
                        reasons.append("QC미달_태도")
                    if "상담" in under_type:
                        reasons.append("QC미달_상담")
                underperforming_reason = ",".join(reasons) if reasons else "QC미달"

                records.append({
                    "record_id": make_id("광주", eng_id, report_week, sheet_name),
                    "center": "광주",
                    "report_week": report_week,
                    "report_week_label": week_label,
                    "report_date": report_date,
                    "agent_name": name,
                    "agent_english_id": eng_id,
                    "vertical": vertical,
                    "service": service,
                    "channel": channel,
                    "hire_date": hire_date.isoformat() if hire_date else None,
                    "tenure_months": tenure,
                    "qc_attitude_error": att_error,
                    "qc_ops_error": ops_error,
                    "is_underperforming": True,
                    "underperforming_reason": underperforming_reason,
                    "note": under_type,
                    "source_file": filename,
                    "source_post_id": None,
                })

    wb.close()
    return records


def parse_underperforming():
    """용산 + 광주 부진상담사 통합 파싱"""
    print("\n" + "=" * 60)
    print("Phase 3: 부진상담사 → underperforming_weekly")
    print("=" * 60)

    all_records = []

    for center in ["용산", "광주"]:
        print(f"\n  [{center}]")
        excel_files = find_underperforming_excels(center)
        print(f"    전체 엑셀: {len(excel_files)}개")

        selected = select_month_end_files(excel_files, center)
        print(f"    선택(월말): {len(selected)}개")

        for target_month, last_week, fn, fp in selected:
            clean_path = None
            tmpdir = None
            try:
                fn_norm = unicodedata.normalize('NFC', fn)

                if center == "용산":
                    # 용산: calamine 엔진으로 직접 읽기 (autoFilter 정리 불필요)
                    records = parse_yongsan_xlsx(fp, fn_norm, target_month=target_month)
                else:
                    # 광주: openpyxl 사용 → autoFilter XML 정리 필요
                    clean_path, tmpdir = clean_xlsx_autofilter(fp)
                    records = parse_gwangju_xlsx(clean_path, fn_norm, target_month=target_month)

                all_records.extend(records)
                print(f"      {fn_norm}: {len(records)}건 (저품질)")
            except Exception as e:
                print(f"      [ERROR] {fn}: {e}")
                import traceback
                traceback.print_exc()
            finally:
                if tmpdir:
                    shutil.rmtree(tmpdir, ignore_errors=True)

    # 중복 체크
    seen = set()
    deduped = []
    for r in all_records:
        key = (r["center"], r["agent_english_id"], r["report_week_label"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    dup_count = len(all_records) - len(deduped)
    if dup_count:
        print(f"\n  중복 제거: {dup_count}건 → {len(deduped)}건")

    print(f"\n  총: {len(deduped)}건")
    write_ndjson(OUT_DIR / "underperforming_weekly.ndjson", deduped)
    return deduped


# ============================================================
# Phase 4: 교육/테스트 이력
# ============================================================
def parse_education_records(agit_posts):
    print("\n" + "=" * 60)
    print("Phase 4: 교육/테스트 → education_records")
    print("=" * 60)

    edu_cats = {"refresher", "new_hire", "job_test", "qa_eval", "advanced_training"}
    records = []

    edu_parents = {}
    for p in agit_posts:
        if p["post_type"] == "원글" and p["category"] in edu_cats:
            edu_parents[p["post_id"]] = p

    print(f"  교육 원글: {len(edu_parents)}건")

    for pid, p in edu_parents.items():
        content = p["content"]
        service = "기타"
        if "택시" in content:
            service = "택시"
        elif "퀵" in content:
            service = "퀵"
        elif "대리" in content:
            service = "대리"

        record_date = None
        if p["created_at"]:
            try:
                dt = datetime.fromisoformat(str(p["created_at"]).replace("Z", ""))
                record_date = dt.strftime("%Y-%m-%d")
            except:
                pass

        title = content.split("\n")[0][:200] if content else ""
        records.append({
            "record_id": make_id(pid, "edu"),
            "post_id": pid,
            "parent_id": p["parent_id"],
            "center": p["center"],
            "record_type": p["category"],
            "record_date": record_date,
            "service": service,
            "title": title,
            "description": None,
            "author_id": p["author_id"],
            "attachment_url": p["attachment_url"],
            "created_at": p["created_at"],
        })

    edu_comments = 0
    for p in agit_posts:
        if p["post_type"] == "댓글" and p["parent_id"] in edu_parents:
            content = p["content"]
            if not content or len(content) < 20:
                continue
            parent = edu_parents[p["parent_id"]]
            service = "기타"
            if "택시" in content:
                service = "택시"
            elif "퀵" in content:
                service = "퀵"
            elif "대리" in content:
                service = "대리"
            record_date = None
            if p["created_at"]:
                try:
                    dt = datetime.fromisoformat(str(p["created_at"]).replace("Z", ""))
                    record_date = dt.strftime("%Y-%m-%d")
                except:
                    pass
            records.append({
                "record_id": make_id(p["post_id"], "edu_c"),
                "post_id": p["post_id"],
                "parent_id": p["parent_id"],
                "center": p["center"],
                "record_type": parent["category"],
                "record_date": record_date,
                "service": service,
                "title": parent["content"].split("\n")[0][:200] if parent["content"] else "",
                "description": content[:10000],
                "author_id": p["author_id"],
                "attachment_url": p["attachment_url"],
                "created_at": p["created_at"],
            })
            edu_comments += 1

    print(f"  교육 댓글: {edu_comments}건")
    print(f"  총: {len(records)}건")
    write_ndjson(OUT_DIR / "education_records.ndjson", records)
    return records


# ============================================================
# 메인
# ============================================================
def main():
    print("=" * 60)
    print("아지트 품질관리 데이터 통합 파싱 v2")
    print(f"소스: {AGIT_DIR}")
    print(f"시간: {datetime.now()}")
    print("=" * 60)

    agit = parse_agit_posts()
    actions = parse_qc_weekly_actions(agit)
    under = parse_underperforming()
    edu = parse_education_records(agit)

    print("\n" + "=" * 60)
    print("파싱 완료 요약")
    print("=" * 60)
    print(f"  agit_posts:             {len(agit)}건")
    print(f"  qc_weekly_actions:      {len(actions)}건")
    print(f"  underperforming_weekly: {len(under)}건")
    print(f"  education_records:      {len(edu)}건")
    print(f"\nNDJSON 출력: {OUT_DIR}")
    print("\nbq load 명령:")
    for tbl in ["agit_posts", "qc_weekly_actions", "underperforming_weekly", "education_records"]:
        print(f"  bq load --source_format=NEWLINE_DELIMITED_JSON --replace {DATASET_ID}.{tbl} {OUT_DIR}/{tbl}.ndjson")


if __name__ == "__main__":
    main()
